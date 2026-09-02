import io
from datetime import datetime, timezone

import pandas as pd
import streamlit as st

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

from database import (
    finalizar_votacao,
    importar_pessoas,
    obter_configuracao,
    obter_estatisticas,
    obter_historico,
    obter_pessoas,
    registrar_voto,
    reabrir_votacao,
    votacao_ativa,
)

from utils.cracha import leitor_cracha


# ============================================================
# CONFIGURAÇÃO
# ============================================================

st.set_page_config(
    page_title="Contabilização de Votos",
    page_icon="🗳️",
    layout="wide",
)


# ============================================================
# FUNÇÕES AUXILIARES
# ============================================================

def agora_manaus():
    """
    Retorna a data/hora atual em UTC.
    """
    return datetime.now(timezone.utc)


def limpar_estado_votacao():
    """
    Limpa o resultado da última leitura.
    """
    st.session_state.pop(
        "resultado_voto",
        None
    )

    st.session_state.pop(
        "dados_voto",
        None
    )


def limpar_valor_excel(valor):
    """
    Normaliza valores provenientes do Excel.
    """

    if valor is None:
        return ""

    try:
        if pd.isna(valor):
            return ""
    except Exception:
        pass

    if isinstance(valor, int):
        return str(valor).strip()

    if isinstance(valor, float):

        if valor.is_integer():
            return str(int(valor))

        return str(valor).strip()

    valor = str(valor).strip()

    if not valor:
        return ""

    if valor.lower() in {
        "nan",
        "none",
        "null",
    }:
        return ""

    if valor.endswith(".0"):

        parte = valor[:-2]

        if parte.isdigit():
            valor = parte

    valor = valor.replace(
        "\u00a0",
        " "
    )

    return valor.strip()


def preparar_dataframe_streamlit(df):
    """
    Prepara DataFrame para o Streamlit.

    Evita o erro do Arrow quando uma mesma coluna
    possui tipos incompatíveis, como lista e texto.
    """

    if df is None:
        return pd.DataFrame()

    if not isinstance(
        df,
        pd.DataFrame
    ):

        df = pd.DataFrame(df)

    df = df.copy()

    for coluna in df.columns:

        def converter(valor):

            if valor is None:
                return ""

            if isinstance(
                valor,
                (
                    list,
                    dict,
                    tuple,
                    set,
                )
            ):
                return str(valor)

            try:

                if pd.isna(valor):
                    return ""

            except Exception:
                pass

            return valor

        df[coluna] = df[coluna].apply(
            converter
        )

    return df


def preparar_dataframe_excel(df):
    """
    Prepara um DataFrame para gravação no XLSX.

    O openpyxl trabalha melhor quando cada célula
    contém um valor escalar simples.
    """

    if df is None:
        return pd.DataFrame()

    if not isinstance(
        df,
        pd.DataFrame
    ):

        df = pd.DataFrame(df)

    df = df.copy()

    for coluna in df.columns:

        def converter(valor):

            if valor is None:
                return ""

            try:

                if pd.isna(valor):
                    return ""

            except Exception:
                pass

            if isinstance(
                valor,
                (
                    list,
                    dict,
                    tuple,
                    set,
                )
            ):
                return str(valor)

            if isinstance(
                valor,
                (
                    pd.Timestamp,
                    datetime,
                )
            ):
                return valor

            if isinstance(
                valor,
                (
                    str,
                    int,
                    float,
                    bool,
                )
            ):
                return valor

            return str(valor)

        df[coluna] = df[coluna].apply(
            converter
        )

    return df


def ajustar_largura_colunas(ws):
    """
    Ajusta largura das colunas da planilha.
    """

    for coluna in ws.columns:

        maior = 0

        letra = coluna[0].column_letter

        for celula in coluna:

            try:

                valor = celula.value

                if valor is None:
                    continue

                tamanho = len(
                    str(valor)
                )

                if tamanho > maior:
                    maior = tamanho

            except Exception:
                continue

        largura = min(
            maior + 2,
            60
        )

        if largura < 10:
            largura = 10

        ws.column_dimensions[
            letra
        ].width = largura


def formatar_planilha(ws):
    """
    Formatação básica da planilha.
    """

    for celula in ws[1]:

        celula.font = Font(
            bold=True
        )

        celula.fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7"
        )

        celula.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    ws.freeze_panes = "A2"

    ajustar_largura_colunas(ws)


def adicionar_dataframe_ao_excel(
    ws,
    df,
    nome_tabela=None
):
    """
    Adiciona um DataFrame em uma worksheet.
    """

    df = preparar_dataframe_excel(
        df
    )

    if df.empty:

        ws.append(
            ["Nenhum registro"]
        )

        return

    for linha in dataframe_to_rows(
        df,
        index=False,
        header=True
    ):

        ws.append(
            linha
        )

    formatar_planilha(
        ws
    )

    # --------------------------------------------------------
    # TABELA EXCEL
    # --------------------------------------------------------

    if (
        nome_tabela
        and ws.max_row >= 2
        and ws.max_column >= 1
    ):

        ultima_coluna = ws.max_column
        ultima_linha = ws.max_row

        referencia = (
            f"A1:"
            f"{ws.cell(
                row=1,
                column=ultima_coluna
            ).column_letter}"
            f"{ultima_linha}"
        )

        try:

            tabela = Table(
                displayName=nome_tabela,
                ref=referencia
            )

            estilo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )

            tabela.tableStyleInfo = estilo

            ws.add_table(
                tabela
            )

        except Exception:
            pass


def mostrar_estatisticas():
    """
    Exibe os KPIs da votação.
    """

    stats = obter_estatisticas()

    if not stats:
        stats = {}

    total = stats.get(
        "total",
        0
    )

    votos = stats.get(
        "votos",
        0
    )

    nao_votaram = stats.get(
        "nao_votaram",
        0
    )

    participacao = stats.get(
        "participacao",
        0
    )

    tentativas = stats.get(
        "tentativas",
        0
    )

    duplicados = stats.get(
        "duplicados",
        0
    )

    nao_encontrados = stats.get(
        "nao_encontrados",
        0
    )

    col1, col2, col3, col4 = st.columns(
        4
    )

    with col1:

        st.metric(
            "👥 Total cadastrados",
            total
        )

    with col2:

        st.metric(
            "🗳️ Votos contabilizados",
            votos
        )

    with col3:

        st.metric(
            "⏳ Não votaram",
            nao_votaram
        )

    with col4:

        st.metric(
            "📈 Participação",
            f"{participacao:.2f}%"
        )

    st.divider()

    col5, col6, col7 = st.columns(
        3
    )

    with col5:

        st.metric(
            "🔎 Total de tentativas",
            tentativas
        )

    with col6:

        st.metric(
            "🚫 Tentativas duplicadas",
            duplicados
        )

    with col7:

        st.metric(
            "⚠️ Não encontrados",
            nao_encontrados
        )

    st.progress(
        min(
            max(
                participacao / 100,
                0
            ),
            1
        )
    )


def processar_voto(
    identificador
):
    """
    Processa a leitura do crachá.
    """

    if not identificador:

        st.warning(
            "Informe ou aproxime um crachá."
        )

        return

    if not votacao_ativa():

        st.error(
            "🔒 A votação está encerrada."
        )

        return

    resultado = registrar_voto(
        identificador
    )

    if not resultado:

        st.error(
            "Não foi possível processar o voto."
        )

        return

    st.session_state[
        "resultado_voto"
    ] = resultado

    st.session_state[
        "dados_voto"
    ] = resultado

    st.rerun()


# ============================================================
# GERAÇÃO DO EXCEL FINAL
# ============================================================

def gerar_excel_final():
    """
    Gera o arquivo Excel final usando openpyxl diretamente.

    Abas criadas:
    - RESUMO
    - RESULTADO
    - HISTÓRICO

    O workbook é criado explicitamente com pelo menos
    uma worksheet visível, eliminando o erro:

    IndexError:
    At least one sheet must be visible
    """

    # --------------------------------------------------------
    # BUSCAR DADOS
    # --------------------------------------------------------

    pessoas = obter_pessoas()

    historico = obter_historico()

    stats = obter_estatisticas()

    config = obter_configuracao()

    # --------------------------------------------------------
    # GARANTIR DATAFRAMES
    # --------------------------------------------------------

    if pessoas is None:
        pessoas = pd.DataFrame()

    if historico is None:
        historico = pd.DataFrame()

    if not isinstance(
        pessoas,
        pd.DataFrame
    ):

        pessoas = pd.DataFrame(
            pessoas
        )

    if not isinstance(
        historico,
        pd.DataFrame
    ):

        historico = pd.DataFrame(
            historico
        )

    # --------------------------------------------------------
    # NORMALIZAR
    # --------------------------------------------------------

    pessoas = preparar_dataframe_excel(
        pessoas
    )

    historico = preparar_dataframe_excel(
        historico
    )

    # --------------------------------------------------------
    # CONFIGURAÇÃO
    # --------------------------------------------------------

    data_finalizacao = ""

    if isinstance(
        config,
        dict
    ):

        data_finalizacao = config.get(
            "data_finalizacao",
            ""
        )

    elif config is not None:

        try:

            data_finalizacao = str(
                config
            )

        except Exception:

            data_finalizacao = ""

    # --------------------------------------------------------
    # RESUMO
    # --------------------------------------------------------

    if not stats:
        stats = {}

    resumo = pd.DataFrame(
        {
            "Indicador": [
                "Total de cadastrados",
                "Votos contabilizados",
                "Não votaram",
                "Participação (%)",
                "Total de tentativas",
                "Votos duplicados",
                "Não encontrados",
                "Data de finalização",
            ],
            "Valor": [
                stats.get(
                    "total",
                    0
                ),
                stats.get(
                    "votos",
                    0
                ),
                stats.get(
                    "nao_votaram",
                    0
                ),
                round(
                    stats.get(
                        "participacao",
                        0
                    ),
                    2
                ),
                stats.get(
                    "tentativas",
                    0
                ),
                stats.get(
                    "duplicados",
                    0
                ),
                stats.get(
                    "nao_encontrados",
                    0
                ),
                data_finalizacao,
            ],
        }
    )

    # --------------------------------------------------------
    # CRIAR WORKBOOK EXPLICITAMENTE
    # --------------------------------------------------------

    workbook = Workbook()

    # A planilha padrão já existe.
    resumo_ws = workbook.active

    resumo_ws.title = "RESUMO"

    # Criar explicitamente as demais planilhas.
    resultado_ws = workbook.create_sheet(
        "RESULTADO"
    )

    historico_ws = workbook.create_sheet(
        "HISTÓRICO"
    )

    # --------------------------------------------------------
    # GARANTIR QUE AS TRÊS PLANILHAS ESTÃO VISÍVEIS
    # --------------------------------------------------------

    resumo_ws.sheet_state = "visible"
    resultado_ws.sheet_state = "visible"
    historico_ws.sheet_state = "visible"

    # A primeira planilha será a ativa.
    workbook.active = 0

    # --------------------------------------------------------
    # ESCREVER RESUMO
    # --------------------------------------------------------

    adicionar_dataframe_ao_excel(
        resumo_ws,
        resumo,
        "TabelaResumo"
    )

    # --------------------------------------------------------
    # ESCREVER RESULTADO
    # --------------------------------------------------------

    adicionar_dataframe_ao_excel(
        resultado_ws,
        pessoas,
        "TabelaResultado"
    )

    # --------------------------------------------------------
    # ESCREVER HISTÓRICO
    # --------------------------------------------------------

    adicionar_dataframe_ao_excel(
        historico_ws,
        historico,
        "TabelaHistorico"
    )

    # --------------------------------------------------------
    # GARANTIA FINAL
    # --------------------------------------------------------

    worksheets_visiveis = [
        ws
        for ws in workbook.worksheets
        if ws.sheet_state == "visible"
    ]

    if len(worksheets_visiveis) == 0:

        workbook.worksheets[0].sheet_state = (
            "visible"
        )

        workbook.active = 0

    # --------------------------------------------------------
    # SALVAR NA MEMÓRIA
    # --------------------------------------------------------

    output = io.BytesIO()

    workbook.save(
        output
    )

    output.seek(0)

    return output


# ============================================================
# CABEÇALHO
# ============================================================

st.title(
    "🗳️ Sistema de Contabilização de Votos"
)

st.caption(
    "Sistema de votação integrado ao Supabase"
)


# ============================================================
# STATUS
# ============================================================

config = obter_configuracao()

ativa = votacao_ativa()

if ativa:

    st.success(
        "🟢 VOTAÇÃO ATIVA"
    )

else:

    st.error(
        "🔴 VOTAÇÃO ENCERRADA"
    )


# ============================================================
# ABAS
# ============================================================

(
    aba_votacao,
    aba_dashboard,
    aba_historico,
    aba_importacao,
) = st.tabs(
    [
        "🗳️ VOTAÇÃO",
        "📊 DASHBOARD",
        "📋 HISTÓRICO",
        "📥 IMPORTAR CADASTRO",
    ]
)


# ============================================================
# ABA VOTAÇÃO
# ============================================================

with aba_votacao:

    st.header(
        "🗳️ Contabilização de Voto"
    )

    if not ativa:

        st.warning(
            "🔒 A votação foi encerrada. "
            "Não é possível registrar novos votos."
        )

    else:

        st.info(
            "Aproxime o crachá do leitor "
            "ou digite o identificador manualmente."
        )

        identificador = st.text_input(
            "Número do crachá",
            key="campo_cracha",
            placeholder=(
                "Aproxime o crachá ou digite o número"
            ),
        )

        # ----------------------------------------------------
        # LEITOR
        # ----------------------------------------------------

        leitor_cracha()

        # ----------------------------------------------------
        # BOTÃO MANUAL
        # ----------------------------------------------------

        col1, col2, col3 = st.columns(
            [1, 1, 1]
        )

        with col2:

            registrar = st.button(
                "🗳️ REGISTRAR VOTO",
                use_container_width=True,
                type="primary",
                key="registrar_voto",
            )

        if registrar:

            processar_voto(
                identificador
            )

        # ----------------------------------------------------
        # RESULTADO
        # ----------------------------------------------------

        if (
            "resultado_voto"
            in st.session_state
        ):

            resultado = (
                st.session_state[
                    "resultado_voto"
                ]
            )

            tipo = resultado.get(
                "resultado",
                ""
            )

            st.divider()

            if tipo == "contabilizado":

                st.success(
                    "✅ VOTO CONTABILIZADO"
                )

                st.markdown(
                    f"""
### 👤 Colaborador

**Nome:** {resultado.get("nome", "")}

**ID Mat:** {resultado.get("id_mat", "")}

**Matrícula:** {resultado.get("mat", "")}

**Voto:** {resultado.get("voto", "SIM")}
"""
                )

            elif tipo == "duplicado":

                st.error(
                    "🚫 VOTO DUPLICADO"
                )

                st.markdown(
                    f"""
### ⚠️ Atenção

O colaborador abaixo **já havia realizado o voto**.

**Nome:** {resultado.get("nome", "")}

**ID Mat:** {resultado.get("id_mat", "")}

**Matrícula:** {resultado.get("mat", "")}

**Voto registrado:** {resultado.get("voto", "SIM")}
"""
                )

            elif tipo == "nao_encontrado":

                identificador_resultado = (
                    resultado.get(
                        "identificador",
                        identificador
                    )
                )

                st.warning(
                    "⚠️ COLABORADOR NÃO ENCONTRADO"
                )

                st.markdown(
                    f"""
O identificador **{identificador_resultado}**
não foi encontrado no cadastro.
"""
                )

            elif tipo == "encerrada":

                st.error(
                    "🔒 VOTAÇÃO ENCERRADA"
                )

            else:

                st.warning(
                    "⚠️ Resultado não identificado."
                )

            # ------------------------------------------------
            # NOVA LEITURA
            # ------------------------------------------------

            if st.button(
                "🔄 NOVA LEITURA",
                use_container_width=True,
                key="nova_leitura",
            ):

                limpar_estado_votacao()

                st.rerun()


# ============================================================
# ABA DASHBOARD
# ============================================================

with aba_dashboard:

    st.header(
        "📊 Dashboard da Votação"
    )

    mostrar_estatisticas()

    st.divider()

    stats = obter_estatisticas()

    if not stats:
        stats = {}

    total = stats.get(
        "total",
        0
    )

    votos = stats.get(
        "votos",
        0
    )

    nao_votaram = stats.get(
        "nao_votaram",
        0
    )

    st.subheader(
        "📌 Resumo"
    )

    col1, col2 = st.columns(
        2
    )

    with col1:

        st.write(
            f"👥 **Total de colaboradores:** {total}"
        )

        st.write(
            f"🗳️ **Votos realizados:** {votos}"
        )

    with col2:

        st.write(
            f"⏳ **Ainda não votaram:** {nao_votaram}"
        )

        st.write(
            f"📈 **Participação:** "
            f"{stats.get('participacao', 0):.2f}%"
        )

    st.divider()

    # --------------------------------------------------------
    # FINALIZAÇÃO
    # --------------------------------------------------------

    st.subheader(
        "🏁 Finalização da Contabilização"
    )

    if ativa:

        st.warning(
            "Atenção: após finalizar a votação, "
            "novos votos não poderão ser registrados."
        )

        confirmar = st.checkbox(
            "Confirmo que desejo finalizar a contabilização.",
            key="confirmar_finalizacao",
        )

        if confirmar:

            if st.button(
                "🏁 FINALIZAR CONTABILIZAÇÃO",
                type="primary",
                use_container_width=True,
                key="finalizar_votacao",
            ):

                resultado_final = (
                    finalizar_votacao()
                )

                if (
                    resultado_final
                    and resultado_final.get(
                        "sucesso"
                    )
                ):

                    st.success(
                        "✅ Contabilização finalizada com sucesso."
                    )

                    st.rerun()

                else:

                    mensagem = (
                        resultado_final.get(
                            "mensagem",
                            "Não foi possível finalizar."
                        )
                        if resultado_final
                        else
                        "Não foi possível finalizar."
                    )

                    st.error(
                        mensagem
                    )

    else:

        st.success(
            "🏁 Contabilização finalizada."
        )

        if config:

            if isinstance(
                config,
                dict
            ):

                data_finalizacao = config.get(
                    "data_finalizacao"
                )

                if data_finalizacao:

                    st.info(
                        f"Finalizada em: "
                        f"{data_finalizacao}"
                    )

        # ----------------------------------------------------
        # GERAR EXCEL FINAL
        # ----------------------------------------------------

        try:

            arquivo_final = gerar_excel_final()

            st.download_button(
                label="📥 BAIXAR PLANILHA FINAL",
                data=arquivo_final,
                file_name=(
                    "resultado_final_votacao.xlsx"
                ),
                mime=(
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
                use_container_width=True,
                key="baixar_planilha_final",
            )

        except Exception as erro:

            st.error(
                "❌ Não foi possível gerar a planilha final."
            )

            st.exception(
                erro
            )

        st.divider()

        # ----------------------------------------------------
        # REABRIR
        # ----------------------------------------------------

        st.subheader(
            "🔓 Reabrir votação"
        )

        st.warning(
            "Use esta opção somente se precisar corrigir "
            "ou continuar a contabilização."
        )

        if st.button(
            "🔓 REABRIR VOTAÇÃO",
            use_container_width=True,
            key="reabrir_votacao",
        ):

            sucesso = reabrir_votacao()

            if sucesso:

                st.success(
                    "Votação reaberta."
                )

                st.rerun()

            else:

                st.error(
                    "Não foi possível reabrir a votação."
                )


# ============================================================
# ABA HISTÓRICO
# ============================================================

with aba_historico:

    st.header(
        "📋 Histórico das Tentativas"
    )

    historico = obter_historico()

    if (
        historico is None
        or len(historico) == 0
    ):

        st.info(
            "Nenhuma tentativa registrada."
        )

    else:

        historico = preparar_dataframe_streamlit(
            historico
        )

        # ----------------------------------------------------
        # FILTROS
        # ----------------------------------------------------

        col1, col2 = st.columns(
            2
        )

        with col1:

            pesquisa = st.text_input(
                "🔎 Pesquisar",
                placeholder=(
                    "Identificador, matrícula, nome ou resultado"
                ),
                key="pesquisa_historico",
            )

        with col2:

            resultados_disponiveis = [
                "TODOS"
            ]

            if "resultado" in historico.columns:

                valores_resultado = (
                    historico[
                        "resultado"
                    ]
                    .fillna("")
                    .astype(str)
                    .unique()
                    .tolist()
                )

                resultados_disponiveis += sorted(
                    valores_resultado
                )

            filtro_resultado = st.selectbox(
                "📌 Resultado",
                resultados_disponiveis,
                key="filtro_resultado",
            )

        df_filtrado = historico.copy()

        # ----------------------------------------------------
        # PESQUISA
        # ----------------------------------------------------

        if pesquisa:

            pesquisa_lower = (
                pesquisa.lower()
            )

            mascara = pd.Series(
                False,
                index=df_filtrado.index,
                dtype=bool,
            )

            for coluna in [
                "identificador",
                "id_mat",
                "mat",
                "nome",
                "resultado",
            ]:

                if coluna in df_filtrado.columns:

                    valores = (
                        df_filtrado[
                            coluna
                        ]
                        .fillna("")
                        .astype(str)
                        .str.lower()
                        .str.contains(
                            pesquisa_lower,
                            regex=False,
                        )
                    )

                    mascara = (
                        mascara
                        |
                        valores
                    )

            df_filtrado = df_filtrado[
                mascara
            ].copy()

        # ----------------------------------------------------
        # FILTRO RESULTADO
        # ----------------------------------------------------

        if (
            filtro_resultado != "TODOS"
            and "resultado"
            in df_filtrado.columns
        ):

            df_filtrado = df_filtrado[
                df_filtrado[
                    "resultado"
                ]
                .astype(str)
                ==
                filtro_resultado
            ].copy()

        # ----------------------------------------------------
        # PREPARAR ARROW
        # ----------------------------------------------------

        df_filtrado_exibicao = (
            preparar_dataframe_streamlit(
                df_filtrado
            )
        )

        # ----------------------------------------------------
        # DATAFRAME
        # ----------------------------------------------------

        st.dataframe(
            df_filtrado_exibicao,
            use_container_width=True,
            hide_index=True,
        )

        # ----------------------------------------------------
        # EXPORTAR HISTÓRICO
        # ----------------------------------------------------

        try:

            historico_excel = (
                preparar_dataframe_excel(
                    df_filtrado
                )
            )

            output_historico = io.BytesIO()

            workbook_historico = Workbook()

            ws_historico = (
                workbook_historico.active
            )

            ws_historico.title = (
                "HISTÓRICO"
            )

            ws_historico.sheet_state = (
                "visible"
            )

            workbook_historico.active = 0

            adicionar_dataframe_ao_excel(
                ws_historico,
                historico_excel,
                "TabelaHistorico"
            )

            workbook_historico.save(
                output_historico
            )

            output_historico.seek(0)

            st.download_button(
                "📥 EXPORTAR HISTÓRICO",
                data=output_historico,
                file_name=(
                    "historico_votacao.xlsx"
                ),
                mime=(
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
                use_container_width=True,
                key="exportar_historico",
            )

        except Exception as erro:

            st.error(
                "❌ Não foi possível gerar o histórico em Excel."
            )

            st.exception(
                erro
            )


# ============================================================
# ABA IMPORTAÇÃO
# ============================================================

with aba_importacao:

    st.header(
        "📥 Importar Cadastro"
    )

    st.info(
        """
O arquivo Excel deve possuir exatamente estas quatro colunas:

**ID Mat | Mat | Nome | Voto**

A coluna **Voto** pode estar vazia para colaboradores
que ainda não votaram.
"""
    )

    st.warning(
        "⚠️ As colunas ID Mat, Mat e Nome são obrigatórias."
    )

    arquivo = st.file_uploader(
        "Selecione o arquivo Excel",
        type=[
            "xlsx",
            "xls",
        ],
        key="arquivo_cadastro",
    )

    if arquivo is not None:

        try:

            # ------------------------------------------------
            # LEITURA
            # ------------------------------------------------

            df = pd.read_excel(
                arquivo,
                header=0,
                dtype=object,
                keep_default_na=False,
                na_filter=False,
            )

            # ------------------------------------------------
            # COLUNAS
            # ------------------------------------------------

            df.columns = [
                str(coluna).strip()
                for coluna in df.columns
            ]

            colunas_esperadas = [
                "ID Mat",
                "Mat",
                "Nome",
                "Voto",
            ]

            # ------------------------------------------------
            # VALIDAR COLUNAS
            # ------------------------------------------------

            if list(df.columns) != (
                colunas_esperadas
            ):

                st.error(
                    "❌ As colunas do arquivo não estão "
                    "no formato esperado."
                )

                st.write(
                    "**Colunas encontradas:**"
                )

                st.code(
                    " | ".join(
                        df.columns.tolist()
                    )
                )

                st.write(
                    "**Colunas esperadas:**"
                )

                st.code(
                    "ID Mat | Mat | Nome | Voto"
                )

                st.stop()

            # ------------------------------------------------
            # LIMPAR
            # ------------------------------------------------

            for coluna in colunas_esperadas:

                df[coluna] = (
                    df[coluna]
                    .apply(
                        limpar_valor_excel
                    )
                )

            # ------------------------------------------------
            # NORMALIZAR VOTO
            # ------------------------------------------------

            df["Voto"] = (
                df["Voto"]
                .astype(str)
                .str.strip()
                .str.upper()
            )

            # ------------------------------------------------
            # REMOVER LINHAS VAZIAS
            # ------------------------------------------------

            df = df[
                ~(
                    (df["ID Mat"] == "")
                    &
                    (df["Mat"] == "")
                    &
                    (df["Nome"] == "")
                    &
                    (df["Voto"] == "")
                )
            ].copy()

            st.success(
                f"✅ Arquivo carregado: "
                f"{len(df)} registros."
            )

            # ------------------------------------------------
            # CAMPOS OBRIGATÓRIOS
            # ------------------------------------------------

            campos_vazios = df[
                (df["ID Mat"] == "")
                |
                (df["Mat"] == "")
                |
                (df["Nome"] == "")
            ].copy()

            # ------------------------------------------------
            # VOTOS INVÁLIDOS
            # ------------------------------------------------

            votos_invalidos = df[
                ~df["Voto"].isin(
                    [
                        "",
                        "SIM",
                    ]
                )
            ].copy()

            # ------------------------------------------------
            # MOSTRAR CAMPOS VAZIOS
            # ------------------------------------------------

            if len(
                campos_vazios
            ) > 0:

                st.error(
                    f"❌ Existem {len(campos_vazios)} "
                    "registro(s) com campos obrigatórios vazios."
                )

                tabela_erros = (
                    preparar_dataframe_streamlit(
                        campos_vazios
                    )
                )

                tabela_erros.insert(
                    0,
                    "Linha Excel",
                    tabela_erros.index + 2,
                )

                tabela_erros = (
                    preparar_dataframe_streamlit(
                        tabela_erros
                    )
                )

                st.dataframe(
                    tabela_erros,
                    use_container_width=True,
                    hide_index=True,
                )

                st.warning(
                    "Corrija essas linhas no Excel "
                    "e envie o arquivo novamente."
                )

            # ------------------------------------------------
            # MOSTRAR VOTOS INVÁLIDOS
            # ------------------------------------------------

            if len(
                votos_invalidos
            ) > 0:

                st.error(
                    f"❌ Existem {len(votos_invalidos)} "
                    "registro(s) com valor inválido na coluna Voto."
                )

                tabela_votos_invalidos = (
                    preparar_dataframe_streamlit(
                        votos_invalidos
                    )
                )

                tabela_votos_invalidos.insert(
                    0,
                    "Linha Excel",
                    tabela_votos_invalidos.index + 2,
                )

                tabela_votos_invalidos = (
                    preparar_dataframe_streamlit(
                        tabela_votos_invalidos
                    )
                )

                st.dataframe(
                    tabela_votos_invalidos,
                    use_container_width=True,
                    hide_index=True,
                )

                st.warning(
                    "A coluna Voto deve conter somente "
                    "SIM ou ficar vazia."
                )

            # ------------------------------------------------
            # PRÉ-VISUALIZAÇÃO
            # ------------------------------------------------

            st.divider()

            st.subheader(
                "👀 Pré-visualização"
            )

            df_visualizacao = (
                preparar_dataframe_streamlit(
                    df.head(20)
                )
            )

            st.dataframe(
                df_visualizacao,
                use_container_width=True,
                hide_index=True,
            )

            # ------------------------------------------------
            # RESUMO
            # ------------------------------------------------

            st.divider()

            st.subheader(
                "📌 Resumo da importação"
            )

            col1, col2, col3, col4 = st.columns(
                4
            )

            with col1:

                st.metric(
                    "Registros",
                    len(df)
                )

            with col2:

                st.metric(
                    "Com voto SIM",
                    int(
                        (
                            df["Voto"]
                            == "SIM"
                        ).sum()
                    )
                )

            with col3:

                st.metric(
                    "Sem voto",
                    int(
                        (
                            df["Voto"]
                            == ""
                        ).sum()
                    )
                )

            with col4:

                st.metric(
                    "Erros",
                    (
                        len(
                            campos_vazios
                        )
                        +
                        len(
                            votos_invalidos
                        )
                    )
                )

            # ------------------------------------------------
            # IMPORTAÇÃO
            # ------------------------------------------------

            if (
                len(campos_vazios) == 0
                and len(votos_invalidos) == 0
                and len(df) > 0
            ):

                st.divider()

                st.success(
                    "✅ Arquivo validado com sucesso. "
                    "Pronto para importação."
                )

                confirmar_importacao = (
                    st.checkbox(
                        "Confirmo que desejo importar este cadastro para o sistema.",
                        key="confirmar_importacao",
                    )
                )

                if confirmar_importacao:

                    if st.button(
                        "📥 IMPORTAR CADASTRO",
                        type="primary",
                        use_container_width=True,
                        key="botao_importar",
                    ):

                        try:

                            with st.spinner(
                                "Importando cadastro para o Supabase..."
                            ):

                                importar_pessoas(
                                    df
                                )

                            st.success(
                                "✅ Cadastro importado com sucesso!"
                            )

                            st.info(
                                "Os dados do Excel agora estão "
                                "disponíveis no sistema."
                            )

                            st.rerun()

                        except Exception as erro:

                            st.error(
                                "❌ Erro durante a importação."
                            )

                            st.exception(
                                erro
                            )

            else:

                if len(df) == 0:

                    st.warning(
                        "⚠️ O arquivo não possui "
                        "registros para importar."
                    )

                elif (
                    len(campos_vazios) > 0
                    or len(votos_invalidos) > 0
                ):

                    st.error(
                        "❌ Corrija os erros acima "
                        "antes de importar."
                    )

        except Exception as erro:

            st.error(
                "❌ Não foi possível ler o arquivo Excel."
            )

            st.exception(
                erro
            )


# ============================================================
# ASSINATURA
# ============================================================

st.markdown(
    """
    <div style="
        text-align: center;
        margin-top: 40px;
        padding: 20px;
        border-top: 2px solid #555;
    ">

            Desenvolvido por Marco Antonio

            Sistema de Contabilização de Votos

    </div>
    """,
    unsafe_allow_html=True
)